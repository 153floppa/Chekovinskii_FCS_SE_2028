"""
Excel ↔ JSON converter for Streamlit integration

Handles Excel uploads, conversion, and JSON export
"""

import openpyxl
import json
from pathlib import Path
from decimal import Decimal
import datetime
from typing import Dict, List, Tuple
import pandas as pd


def convert_to_serializable(obj):
    """Конвертирует объекты в JSON-совместимые типы с максимальной точностью"""
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, datetime.datetime):
        return obj.isoformat()
    elif isinstance(obj, datetime.date):
        return obj.isoformat()
    elif isinstance(obj, datetime.time):
        return obj.isoformat()
    elif hasattr(obj, '__iter__') and not isinstance(obj, (str, bytes)):
        return [convert_to_serializable(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    return obj


def excel_to_dict(excel_file_path: str) -> Dict[str, List[Dict]]:
    """
    Парсит Excel файл и возвращает словарь с листами

    Args:
        excel_file_path: Путь к Excel файлу

    Returns:
        Словарь: {sheet_name: [rows]}
    """
    try:
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []
            headers = []

            # Get headers
            for cell in ws[1]:
                headers.append(cell.value)

            # Get rows
            for row in ws.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    header = headers[col_idx]
                    value = cell.value

                    if header is not None:
                        row_data[header] = convert_to_serializable(value)

                # Skip empty rows
                if any(row_data.values()):
                    sheet_data.append(row_data)

            data[sheet_name] = sheet_data

        wb.close()
        return data

    except Exception as e:
        raise Exception(f"Failed to parse Excel: {str(e)}")


def excel_to_json_file(excel_file_path: str, output_json_path: str) -> Tuple[bool, str]:
    """
    Парсит Excel файл и сохраняет в JSON

    Args:
        excel_file_path: Путь к Excel файлу
        output_json_path: Путь для сохранения JSON

    Returns:
        (success, message)
    """
    try:
        data = excel_to_dict(excel_file_path)

        # Ensure directory exists
        output_path = Path(output_json_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save to JSON
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        # Count rows
        total_rows = sum(len(rows) for rows in data.values())
        sheets_count = len(data)

        message = f"✅ Успешно конвертировано!\n"
        message += f"📄 Листов: {sheets_count}\n"
        message += f"📊 Строк: {total_rows}\n"
        message += f"💾 Сохранено: {output_json_path}"

        return True, message

    except Exception as e:
        return False, f"❌ Ошибка: {str(e)}"


def excel_to_dataframe(excel_file_path: str, sheet_name: int = 0) -> pd.DataFrame:
    """
    Загружает Excel в pandas DataFrame

    Args:
        excel_file_path: Путь к Excel файлу
        sheet_name: Номер листа (0-indexed)

    Returns:
        DataFrame
    """
    try:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        return df
    except Exception as e:
        raise Exception(f"Failed to read Excel: {str(e)}")


def merge_excel_with_existing(
    new_data: List[Dict],
    existing_json_path: str = None
) -> List[Dict]:
    """
    Объединяет данные из новой выгрузки с существующими

    Args:
        new_data: Новые строки из Excel
        existing_json_path: Путь к существующему JSON (опционально)

    Returns:
        Объединённый список
    """
    try:
        if existing_json_path and Path(existing_json_path).exists():
            with open(existing_json_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
                if isinstance(existing_data, dict):
                    # Если это dict со списками листов
                    existing_list = []
                    for sheet_data in existing_data.values():
                        existing_list.extend(sheet_data)
                else:
                    existing_list = existing_data

            # Merge
            merged = existing_list + new_data
            return merged
        else:
            return new_data

    except Exception as e:
        raise Exception(f"Failed to merge data: {str(e)}")
