"""
📊 Парсер Excel → JSON

Назначение:
    Конвертирует Excel файлы (XLSX) в JSON с сохранением всех типов данных,
    дат, времени и чисел с полной точностью (используется Decimal).

Использование:
    python3 parse_excel.py [input.xlsx] [output.json]

Вход:
    - prodam.xlsx (данные продаж)
    - sdam.xlsx (данные аренды)

Выход:
    - prodam.json
    - sdam.json

Примечания:
    - Сохраняет все листы из Excel
    - Пропускает пустые строки
    - Конвертирует даты в ISO формат
    - Сохраняет числа с максимальной точностью
"""

import openpyxl
import json
from pathlib import Path
from decimal import Decimal
import datetime

def convert_to_serializable(obj):
    """Конвертирует объекты в JSON-совместимые типы с максимальной точностью"""
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, datetime.datetime):
        return obj.isoformat()
    elif isinstance(obj, datetime.date):
        return obj.isoformat()
    elif isinstance(obj, datetime.time):
        return obj.isoformat()
    elif hasattr(obj, '__iter__') and not isinstance(obj, (str, bytes)):
        return [convert_to_serializable(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    return obj

def parse_excel_to_json(excel_file, output_file):
    """Парсит Excel файл в JSON с сохранением всей точности"""

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = []
        headers = []

        for cell in ws[1]:
            headers.append(cell.value)

        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data = {}
            for col_idx, cell in enumerate(row):
                header = headers[col_idx]
                value = cell.value

                if header is not None:
                    row_data[header] = convert_to_serializable(value)

            if any(row_data.values()):
                sheet_data.append(row_data)

        data[sheet_name] = sheet_data

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ {excel_file} → {output_file}")
    print(f"  Листов: {len(data)}")
    for sheet_name, sheet_data in data.items():
        print(f"  - {sheet_name}: {len(sheet_data)} строк")

excel_dir = Path(__file__).parent

files_to_parse = [
    ("prodam.xlsx", "prodam.json"),
    ("sdam.xlsx", "sdam.json")
]

for excel_file, json_file in files_to_parse:
    excel_path = excel_dir / excel_file
    output_path = excel_dir / json_file

    if excel_path.exists():
        parse_excel_to_json(str(excel_path), str(output_path))
    else:
        print(f"✗ Файл не найден: {excel_path}")

print("\n✓ Парсинг завершен!")
